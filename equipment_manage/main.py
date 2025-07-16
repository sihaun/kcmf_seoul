import tkinter as tk
from tkinter import filedialog, ttk
import sys
import threading
import tempfile
import json
import os
from utils import paste_equipment, xls2xlsx
from openpyxl import load_workbook

class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

# 메인 앱 클래스
class EquipmentApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("2025년 장비운영실적")
        self.geometry("600x400")

        self.source_path = ""
        self.target_path = ""

        # 프레임 정의
        self.main_frame = MainFrame(self)
        self.reset_frame = ResetFrame(self)

        # stdout 연결
        sys.stdout = StdoutRedirector(self.main_frame.output_text)
        sys.stderr = sys.stdout

        self.show_main()

    def show_main(self):
        self.reset_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)
        sys.stdout = StdoutRedirector(self.main_frame.output_text)

    def show_reset(self):
        self.main_frame.pack_forget()
        self.reset_frame.pack(fill="both", expand=True)
        sys.stdout = StdoutRedirector(self.reset_frame.output_text)

class MainFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # 파일 선택 영역
        frame_top = tk.Frame(self)
        frame_top.pack(pady=10)

        # 담당자 이름 입력
        frame_name = tk.Frame(self)
        frame_name.pack(pady=5)
        tk.Label(frame_name, text="담당자 이름:").pack(side=tk.LEFT)
        self.name_var = tk.StringVar()
        tk.Entry(frame_name, textvariable=self.name_var, width=20).pack(side=tk.LEFT)

        # settings.json 불러오기
        self.settings_path = "config/settings.json"
        try:
            with open(self.settings_path, "r", encoding="utf-8") as f:
                saved = json.load(f)
                self.name_var.set(saved.get("manager_name", ""))
                self.master.source_path = saved.get("source_path", "")
                self.master.target_path = saved.get("target_path", "")
        except:
            pass

        self.source_label = self.create_file_selector(frame_top, "장비운영실적 파일 선택", self.select_source)
        self.source_label.config(text=self.master.source_path or "(경로 없음)")

        self.target_label = self.create_file_selector(frame_top, "양식 파일 선택", self.select_target)
        self.target_label.config(text=self.master.target_path or "(경로 없음)")


        # 처리 시작 버튼
        tk.Button(self, text="처리 시작", command=self.start_process).pack(pady=5)

        # 재설정 이동 버튼
        tk.Button(self, text="재설정 →", command=master.show_reset).pack(pady=5)

        # 진행 바 + 라벨
        self.progress_bar = ttk.Progressbar(self, length=550, mode='determinate')
        self.progress_bar.pack(pady=5)
        self.progress_label = tk.Label(self, text="0% 완료")
        self.progress_label.pack()

        # 출력창
        self.output_text = tk.Text(self, height=10)
        self.output_text.pack(fill="both", expand=True, padx=10, pady=10)

    def save_settings(self):
        os.makedirs("config", exist_ok=True)
        data = {
            "manager_name": self.name_var.get(),
            "source_path": self.master.source_path,
            "target_path": self.master.target_path
        }
        # equipment_path는 ResetFrame 쪽에서 저장함
        try:
            with open(self.settings_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
                if "equipment_path" in existing:
                    data["equipment_path"] = existing["equipment_path"]
        except:
            pass

        with open(self.settings_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


    def create_file_selector(self, parent, text, command):
        frame = tk.Frame(parent)
        frame.pack(side=tk.LEFT, padx=20)
        tk.Button(frame, text=text, command=command).pack()
        label = tk.Label(frame, text="(경로 없음)", wraplength=250)
        label.pack()
        return label

    def select_source(self):
        path = filedialog.askopenfilename(title="장비운영실적 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            if path.endswith(".xls"):
                print("xls 파일 감지됨. xlsx로 변환 중...")
                temp_dir = tempfile.gettempdir()
                xlsx_path = os.path.join(temp_dir, os.path.basename(path).replace(".xls", ".xlsx"))
                xls2xlsx(path, xlsx_path)
                print(f"변환 완료: {xlsx_path}")
                path = xlsx_path
            self.source_label.config(text=path)
            self.master.source_path = path
            self.save_settings()

    def select_target(self):
        path = filedialog.askopenfilename(title="양식 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            if path.endswith(".xls"):
                print("xls 파일 감지됨. xlsx로 변환 중...")
                temp_dir = tempfile.gettempdir()
                xlsx_path = os.path.join(temp_dir, os.path.basename(path).replace(".xls", ".xlsx"))
                xls2xlsx(path, xlsx_path)
                print(f"변환 완료: {xlsx_path}")
                path = xlsx_path
            self.target_label.config(text=path)
            self.master.target_path = path
            self.save_settings()

    def start_process(self):
        if not self.master.source_path or not self.master.target_path:
            print("장비운영실적 파일과 양식 파일을 모두 선택하세요.")
            return

        # 이름 저장
        manager_name = self.name_var.get()
        self.master.manager_name = manager_name
        os.makedirs("config", exist_ok=True)
        with open("config/settings.json", "w", encoding="utf-8") as f:
            json.dump({"manager_name": manager_name}, f, ensure_ascii=False, indent=2)

        threading.Thread(target=self.execute_process_thread, daemon=True).start()

    def execute_process_thread(self):
        source_path = self.master.source_path
        target_path = self.master.target_path
        print(f"장비운영실적.xlsx: {source_path}")
        print(f"양식 파일.xlsx: {target_path}")
        print("처리 시작...")

        wb_source = load_workbook(source_path)
        ws_source = wb_source.active
        wb_target = load_workbook(target_path)
        ws_target = wb_target["7.장비"]

        row = 5
        cnt = 0
        while ws_source[f"E{row}"].value is not None:
            cnt += 1
            row += 1

        with open("config/equipment.json", "r", encoding="utf-8") as f:
            try:
                config_data = json.load(f)
                target_pointer = 4
                manager_name = self.master.manager_name
                for source_pointer in range(5, cnt + 5):
                    paste_equipment(manager_name, config_data, ws_source, ws_target, source_pointer, target_pointer)
                    target_pointer += 1

                    rate = (source_pointer - 4) * 100 // cnt
                    self.progress_bar["value"] = rate
                    self.progress_label.config(text=f"{rate}% 완료")
                    self.update_idletasks()

                wb_target.save(target_path)
                print("처리 완료!")
            except:
                print("장비 종류 파일이 설정되지 않았습니다. '재설정' 으로 이동하여 장비 파일을 설정해주세요.")

class ResetFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # 파일 선택
        tk.Button(self, text="장비종류xlsx 선택", command=self.select_file).pack(pady=5)
        self.file_label = tk.Label(self, text="(파일 없음)")
        self.file_label.pack()

        try:
            with open("config/settings.json", "r", encoding="utf-8") as f:
                saved = json.load(f)
                equipment_path = saved.get("equipment_path", "")
                if equipment_path:
                    self.file_label.config(text=equipment_path)
                    self.source_path = equipment_path
        except:
            pass

        # 실행 버튼
        tk.Button(self, text="재설정", command=self.start_reset).pack(pady=5)

        # 뒤로가기
        tk.Button(self, text="← 뒤로가기", command=master.show_main).pack(pady=5)

        # 진행 바 + 라벨
        self.progress_bar = ttk.Progressbar(self, length=550, mode='determinate')
        self.progress_bar.pack(pady=5)
        self.progress_label = tk.Label(self, text="0% 완료")
        self.progress_label.pack()

        # 출력창
        self.output_text = tk.Text(self, height=10)
        self.output_text.pack(fill="both", expand=True, padx=10, pady=10)

    def select_file(self):
        path = filedialog.askopenfilename(title="장비 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            if path.endswith(".xls"):
                print("xls 파일 감지됨. xlsx로 변환 중...")
                temp_dir = tempfile.gettempdir()
                xlsx_path = os.path.join(temp_dir, os.path.basename(path).replace(".xls", ".xlsx"))
                xls2xlsx(path, xlsx_path)
                print(f"변환 완료: {xlsx_path}")
                path = xlsx_path
            self.file_label.config(text=path)
            self.source_path = path
            os.makedirs("config", exist_ok=True)

            # 기존 설정 불러와 병합
            settings_path = "config/settings.json"
            data = {"equipment_path": path}
            try:
                with open(settings_path, "r", encoding="utf-8") as f:
                    existing = json.load(f)
                    data.update(existing)
            except:
                pass

            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

    def start_reset(self):
        if not hasattr(self, 'source_path'):
            print("파일을 먼저 선택하세요.")
            return
        threading.Thread(target=self.reset_equipment_thread, args=(self.source_path,), daemon=True).start()

    def reset_equipment_thread(self, path):
        print("처리 시작...")
        wb = load_workbook(path)
        ws = wb.active

        row = 2
        cnt = 0
        while ws[f"A{row}"].value is not None:
            cnt += 1
            row += 1

        data_dict = {}
        for pointer in range(2, cnt + 2):
            if ws[f"C{pointer}"].value is None:
                continue

            key = ws[f"C{pointer}"].value
            equipment_group = ws[f"B{pointer}"].value
            model_name = ws[f"D{pointer}"].value

            match equipment_group:
                case "기타" | "편집":
                    equipment_group = "스마트장비"
                case "녹음" | "녹음보조장비":
                    equipment_group = "녹음장비"
                case "조명" | "조명보조장비":
                    equipment_group = "조명장비"
                case "삼각대" | "촬영보조장비":
                    equipment_group = "촬영보조"
                case "카메라":
                    equipment_group = "촬영장비"

            data_dict[key] = {
                "equipment_group": equipment_group,
                "model_name": model_name,
            }

            rate = (pointer - 1) * 100 // cnt
            self.progress_bar["value"] = rate
            self.progress_label.config(text=f"{rate}% 완료")
            self.update_idletasks()

        with open("config/equipment.json", "w", encoding="utf-8") as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=2)

        print("처리 완료!")

if __name__ == "__main__":
    app = EquipmentApp()
    app.mainloop()
