from openpyxl import load_workbook
#from tqdm import tqdm
import os
from copy import copy

'''
for school in os.listdir(path):
    school_path = os.path.join(path, school)
'''
def return_class(class_name : str, school_name : str) -> str:
    class_map = {"촬영감독" : f"[공통_청소년 온라인 미디어 진로특강]{school_name}(촬영감독)",
              "아나운서" : f"[공통_청소년 온라인 미디어 진로특강]{school_name}(아나운서)",
              "크리에이터" : f"[공통_청소년 온라인 미디어 진로특강]{school_name}(크리에이터)",
              "미디어아트" : f"[공통_청소년 온라인 미디어 진로특강]{school_name}(미디어아트)",
              "미디어리터러시" : f"[공통_청소년 온라인 미디어 진로특강]{school_name}(미디어리터러시 웹드라마)",
              }
    return class_map[class_name]

def listing_student(dir_path : str, school_name : str):
    school_path = os.path.join(dir_path, school_name)


    # 2. 폴더 내 .xlsx 파일 목록 수집
    excel_files = [f for f in os.listdir(school_path) if f.endswith(".xlsx")]

    # 예외 처리: 엑셀 파일이 없을 경우
    if not excel_files:
        print(f"{school_name} 폴더에 .xlsx 파일이 없습니다.")
        raise ValueError(f"{school_name} 폴더에 .xlsx 파일이 없습니다.")

    if len(excel_files) > 1:
        print(f"{school_path} 폴더에 .xlsx 파일이 여러 개 있습니다. 하나만 있어야 합니다.")
        raise FileExistsError(f"{school_path} 폴더에 .xlsx 파일이 여러 개 있습니다. 하나만 있어야 합니다.")

    wb = load_workbook(os.path.join(school_path, excel_files[0]))
    ws = wb.active

    # 값과 일부 서식 복사 (수동으로)
    name_list = []
    gender_list = []
    pointer = 5
    for row in range(pointer, ws.max_row + 1):
        name_list.append(ws[f"B{row}"].value)
        gender_list.append(ws[f"D{row}"].value)

    name_count = {}
    result = []

    for name in name_list:
        # 이름 카운트 증가
        name_count[name] = name_count.get(name, 0) + 1

        # 두 자리 숫자로 포맷
        suffix = f"{name_count[name]:02d}"

        # 이름 + 순번 조합
        result.append(f"{name}{suffix}")

    result_name = [name[:-2] if name.endswith("01") else name for name in result]

    return result_name, gender_list


def perfectcopy(target, source, value=None):
    target.font = copy(source.font)
    target.border = copy(source.border)
    target.fill = copy(source.fill)
    target.number_format = copy(source.number_format)
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)

    if value is None:
        target.value = source.value
    else:
        target.value = value
    
def paste_school(source_sheet, target_sheet, source_pointer: int=2, target_counter : int=7):
    ws_source = source_sheet

    ws_target = target_sheet

    school_name = ws_source[f"B{source_pointer}"].value
    class_name = ws_source[f"C{source_pointer}"].value
    indexing_map = {"촬영감독" : 2,
            "아나운서" : 3,
            "크리에이터" : 4,
            "미디어아트" : 5,
            "미디어리터러시" : 6,
            }
    index = indexing_map[class_name]
    date = ws_source[f"D{source_pointer}"].value
    day = date2day(date)
    start = ws_source[f"E{source_pointer}"].value
    end = ws_source[f"F{source_pointer}"].value
    num_people = ws_source[f"G{source_pointer}"].value

    for i in range(65, 91, 1):
        column = chr(i)
        if column == "K":
            value = return_class(class_name, school_name)
        elif column == "M":
            value = date
        elif column == "N":
            value = day
        elif column == "O":
            value = start
        elif column == "P":
            value = end
        elif column == "R":
            value = num_people
        elif column == "S":
            value = num_people             
        else : 
            value = None
        perfectcopy(ws_target[f"{column}{target_counter}"], ws_target[f"{column}{index}"], value)

    perfectcopy(ws_target[f"AA{target_counter}"], ws_target[f"AA{index}"], school_name)
    perfectcopy(ws_target[f"AB{target_counter}"], ws_target[f"AB{index}"])
    perfectcopy(ws_target[f"AC{target_counter}"], ws_target[f"AC{index}"])

    target_counter += 1

    return target_counter


def paste_student(dir_path, source_sheet, target_sheet, source_pointer: int=2, target_counter : int=4):

    ws_source = source_sheet

    ws_target = target_sheet

    list_counter = 5
    try:
        name_list, gender_list = listing_student(dir_path, ws_source[f"A{source_pointer}"].value)

        for i in range(ws_source[f"G{source_pointer}"].value): # source의 인구수만큼 반복
            school_name = ws_source[f"B{source_pointer}"].value
            class_name = ws_source[f"C{source_pointer}"].value
            try:
                name = name_list[i]
                gender = gender_list[i]
            except:
                print(f"school : {school_name}, index : {list_counter}. 인원과 명단 수가 맞지 않음.")
            address = ws_source[f"H{source_pointer}"].value

            if name == None:
                print(f"school : {school_name}, index : {list_counter}. 인원과 명단 수가 맞지 않음.")
                raise IndexError(f"index : {list_counter}. 인원과 명단 수가 맞지 않음.")
            # 80
            for j in range(65, 80, 1):
                column = chr(j)
                if column == "D":
                    value = return_class(class_name, school_name)
                elif column == "G":
                    value = name
                elif column == "J":
                    value = "100%"
                elif column == "L":
                    value = gender
                elif column == "O":
                    value = address
                else : 
                    value = None
                perfectcopy(ws_target[f"{column}{target_counter}"], ws_target[f"{column}2"], value)

            target_counter += 1
            list_counter += 1

    except:
        print("레전드 상황 발생")

    return target_counter

# date : "2025-06-18"
def date2day(date):
    date_obj = date

    # 한글 요일 매핑
    korean_weekdays = ['월', '화', '수', '목', '금', '토', '일']
    weekday_kor = korean_weekdays[date_obj.weekday()]

    return weekday_kor


def execute_process(dir_path, source_path="temp.xlsx", target_path="main.xlsx"):
    print(f"소스: {source_path}")
    print(f"타겟: {target_path}")
    print("처리 시작...")
    wb_source = load_workbook(source_path)
    ws_source = wb_source.active
    
    wb_target = load_workbook(target_path)
    ws_target_school = wb_target["2.교육(일지)"]
    ws_target_student = wb_target["3.교육(명부)"]

    source_pointer = 2
    target_counter_school = 8
    target_counter_student = 5

    visited = ""
    while ws_source[f"A{source_pointer}"].value is not None:
        target_counter_school = paste_school(ws_source, ws_target_school, source_pointer, target_counter_school)

        if visited != ws_source[f"A{source_pointer}"].value:
            target_counter_student = paste_student(dir_path, ws_source, ws_target_student, source_pointer, target_counter_student)
            visited = ws_source[f"A{source_pointer}"].value
        
        source_pointer += 1

    wb_target.save(target_path)
    print("처리 완료!")